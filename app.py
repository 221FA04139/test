
import os
import io
import json
import secrets
from datetime import datetime, date, timedelta
from decimal import Decimal

from dotenv import load_dotenv
load_dotenv()

from flask import (
    Flask, render_template, request, redirect, url_for,
    jsonify, session, send_file, abort,
)
from flask_login import (
    LoginManager, login_user, logout_user, login_required, current_user,
)
from flask_migrate import Migrate
from flask_wtf.csrf import CSRFProtect, generate_csrf

from database import (
    db, User, Transaction, OpeningBalance, Category, AuditLog,
    ACCOUNTS, CATEGORIES, seed_database,
)

# ── App setup ─────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", secrets.token_hex(32))

database_url = os.environ.get("DATABASE_URL", "")
# Railway / Heroku use postgres:// — SQLAlchemy 2.x needs postgresql://
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
if not database_url:
    raise RuntimeError("DATABASE_URL environment variable is not set.")

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=24)
app.config["WTF_CSRF_TIME_LIMIT"] = None  # no expiry on CSRF tokens

db.init_app(app)
migrate = Migrate(app, db)
csrf = CSRFProtect(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = ""


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ── Helpers ───────────────────────────────────────────────────────────────────

def log_audit(action, table_name, record_id, old_value=None, detail=None):
    entry = AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        action=action,
        table_name=table_name,
        record_id=record_id,
        old_value=old_value,
        detail=detail,
    )
    db.session.add(entry)


def get_account_balance(account):
    ob = OpeningBalance.query.filter_by(account=account).first()
    opening = float(ob.balance) if ob else 0.0

    income = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.account == account,
        Transaction.type == "Income",
        Transaction.is_deleted == False,
    ).scalar() or 0

    out = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.account == account,
        Transaction.type.in_(["Expense", "Transfer"]),
        Transaction.is_deleted == False,
    ).scalar() or 0

    transfer_in = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.to_account == account,
        Transaction.type == "Transfer",
        Transaction.is_deleted == False,
    ).scalar() or 0

    return opening + float(income) - float(out) + float(transfer_in)


def _today_str():
    return date.today().isoformat()


def _current_month():
    return date.today().strftime("%Y-%m")


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        remember = request.form.get("remember") == "on"

        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.check_password(password):
            login_user(user, remember=remember, duration=timedelta(days=30))
            session.permanent = True
            return redirect(url_for("index"))
        error = "Invalid username or password."

    return render_template("login.html", error=error, csrf_token=generate_csrf())


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── Main app ──────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    version = os.environ.get("APP_VERSION", "1.0.0")
    return render_template(
        "app.html",
        user=current_user,
        accounts=ACCOUNTS,
        categories=CATEGORIES,
        version=version,
        csrf_token=generate_csrf(),
    )


# ── API: Transactions ─────────────────────────────────────────────────────────

@app.route("/api/transactions", methods=["GET"])
@login_required
def api_transactions():
    page = request.args.get("page", 1, type=int)
    per_page = 50

    q = Transaction.query.filter_by(is_deleted=False)

    account = request.args.get("account")
    tx_type = request.args.get("type")
    category = request.args.get("category")
    month = request.args.get("month")
    search = request.args.get("search", "").strip()

    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")

    if account:
        q = q.filter(Transaction.account == account)
    if tx_type:
        q = q.filter(Transaction.type == tx_type)
    if category:
        q = q.filter(Transaction.category == category)
    if month:
        q = q.filter(Transaction.month == month)
    try:
        if from_date:
            q = q.filter(Transaction.date >= date.fromisoformat(from_date))
        if to_date:
            q = q.filter(Transaction.date <= date.fromisoformat(to_date))
    except ValueError:
        pass
    if search:
        like = f"%{search}%"
        q = q.filter(
            db.or_(
                Transaction.description.ilike(like),
                Transaction.ref.ilike(like),
                Transaction.account.ilike(like),
                Transaction.category.ilike(like),
                Transaction.subcat.ilike(like),
            )
        )

    total = q.count()
    txs = (
        q.order_by(Transaction.date.desc(), Transaction.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return jsonify({
        "transactions": [t.to_dict() for t in txs],
        "total": total,
        "page": page,
        "pages": max(1, (total + per_page - 1) // per_page),
        "per_page": per_page,
    })


@app.route("/api/transactions", methods=["POST"])
@login_required
def api_create_transaction():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    required = ["date", "type", "account", "category", "amount"]
    for field in required:
        if not data.get(field):
            return jsonify({"error": f"Field '{field}' is required"}), 400

    if data["type"] not in ("Income", "Expense", "Transfer"):
        return jsonify({"error": "Invalid transaction type"}), 400

    try:
        tx_date = date.fromisoformat(data["date"])
        amount = Decimal(str(data["amount"]))
        if amount <= 0:
            raise ValueError
    except (ValueError, Exception):
        return jsonify({"error": "Invalid date or amount"}), 400

    to_account = data.get("to_account") or None
    if data["type"] == "Transfer" and to_account and to_account == data["account"]:
        return jsonify({"error": "To Account must differ from Account"}), 400

    tx = Transaction(
        date=tx_date,
        type=data["type"],
        account=data["account"],
        category=data["category"],
        subcat=data.get("subcat") or None,
        description=data.get("description") or None,
        amount=amount,
        ref=data.get("ref") or None,
        to_account=to_account,
        month=tx_date.strftime("%Y-%m"),
        entered_by=current_user.id,
    )
    db.session.add(tx)
    db.session.flush()
    log_audit("CREATE", "transactions", tx.id, detail=json.dumps(tx.to_dict()))
    db.session.commit()

    return jsonify({"transaction": tx.to_dict(), "message": "Transaction saved"}), 201


@app.route("/api/transactions/<int:tx_id>", methods=["PUT"])
@login_required
def api_update_transaction(tx_id):
    tx = Transaction.query.filter_by(id=tx_id, is_deleted=False).first_or_404()
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    old = json.dumps(tx.to_dict())

    try:
        if "date" in data:
            tx.date = date.fromisoformat(data["date"])
            tx.month = tx.date.strftime("%Y-%m")
        if "type" in data:
            if data["type"] not in ("Income", "Expense", "Transfer"):
                return jsonify({"error": "Invalid type"}), 400
            tx.type = data["type"]
        if "account" in data:
            tx.account = data["account"]
        if "category" in data:
            tx.category = data["category"]
        if "subcat" in data:
            tx.subcat = data["subcat"] or None
        if "description" in data:
            tx.description = data["description"] or None
        if "amount" in data:
            amount = Decimal(str(data["amount"]))
            if amount <= 0:
                raise ValueError
            tx.amount = amount
        if "ref" in data:
            tx.ref = data["ref"] or None
        if "to_account" in data:
            tx.to_account = data["to_account"] or None
    except (ValueError, Exception):
        return jsonify({"error": "Invalid data"}), 400

    tx.updated_at = datetime.utcnow()
    log_audit("UPDATE", "transactions", tx.id, old_value=old, detail=json.dumps(tx.to_dict()))
    db.session.commit()

    return jsonify({"transaction": tx.to_dict(), "message": "Transaction updated"})


@app.route("/api/transactions/<int:tx_id>", methods=["DELETE"])
@login_required
def api_delete_transaction(tx_id):
    tx = Transaction.query.filter_by(id=tx_id, is_deleted=False).first_or_404()
    old = json.dumps(tx.to_dict())
    tx.is_deleted = True
    tx.updated_at = datetime.utcnow()
    log_audit("DELETE", "transactions", tx.id, old_value=old)
    db.session.commit()
    return jsonify({"message": "Transaction deleted"})


# ── API: Dashboard ────────────────────────────────────────────────────────────

@app.route("/api/dashboard")
@login_required
def api_dashboard():
    today = date.today()
    today_str = today.isoformat()
    current_month = today.strftime("%Y-%m")

    # All-time totals
    all_income = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == "Income", Transaction.is_deleted == False
    ).scalar() or 0

    all_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == "Expense", Transaction.is_deleted == False
    ).scalar() or 0

    # This month
    month_income = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == "Income",
        Transaction.month == current_month,
        Transaction.is_deleted == False,
    ).scalar() or 0

    month_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
        Transaction.type == "Expense",
        Transaction.month == current_month,
        Transaction.is_deleted == False,
    ).scalar() or 0

    # Today
    today_txs = Transaction.query.filter_by(is_deleted=False).filter(
        Transaction.date == today
    ).all()
    today_income = sum(float(t.amount) for t in today_txs if t.type == "Income")
    today_expense = sum(float(t.amount) for t in today_txs if t.type == "Expense")

    # Account balances
    account_balances = []
    for acc in ACCOUNTS:
        ob = OpeningBalance.query.filter_by(account=acc).first()
        opening = float(ob.balance) if ob else 0.0
        income = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.account == acc,
            Transaction.type == "Income",
            Transaction.is_deleted == False,
        ).scalar() or 0
        out = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.account == acc,
            Transaction.type.in_(["Expense", "Transfer"]),
            Transaction.is_deleted == False,
        ).scalar() or 0
        transfer_in = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.to_account == acc,
            Transaction.type == "Transfer",
            Transaction.is_deleted == False,
        ).scalar() or 0
        balance = opening + float(income) - float(out) + float(transfer_in)
        account_balances.append({
            "account": acc,
            "opening": opening,
            "total_in": float(income) + float(transfer_in),
            "total_out": float(out),
            "balance": balance,
        })

    # Category P&L
    cat_rows = (
        db.session.query(
            Transaction.category,
            Transaction.type,
            db.func.sum(Transaction.amount).label("total"),
        )
        .filter(Transaction.is_deleted == False)
        .group_by(Transaction.category, Transaction.type)
        .all()
    )
    cat_map = {}
    for cat, ttype, total in cat_rows:
        if cat not in cat_map:
            cat_map[cat] = {"category": cat, "income": 0.0, "expense": 0.0}
        if ttype == "Income":
            cat_map[cat]["income"] += float(total)
        elif ttype in ("Expense", "Transfer"):
            cat_map[cat]["expense"] += float(total)
    for v in cat_map.values():
        v["net"] = v["income"] - v["expense"]
    category_pnl = list(cat_map.values())

    # Monthly trend — last 12 months (proper month arithmetic, no timedelta hack)
    monthly = []
    for i in range(11, -1, -1):
        month_num = today.month - i
        year_num = today.year
        while month_num <= 0:
            month_num += 12
            year_num -= 1
        m = f"{year_num:04d}-{month_num:02d}"
        mi = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == "Income",
            Transaction.month == m,
            Transaction.is_deleted == False,
        ).scalar() or 0
        me = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.type == "Expense",
            Transaction.month == m,
            Transaction.is_deleted == False,
        ).scalar() or 0
        monthly.append({"month": m, "income": float(mi), "expense": float(me)})

    # Category donut (expenses only)
    donut_data = [
        {"category": v["category"], "amount": v["expense"]}
        for v in category_pnl if v["expense"] > 0
    ]

    return jsonify({
        "all_time": {
            "income": float(all_income),
            "expense": float(all_expense),
            "net": float(all_income) - float(all_expense),
        },
        "this_month": {
            "income": float(month_income),
            "expense": float(month_expense),
            "net": float(month_income) - float(month_expense),
        },
        "today": {
            "count": len(today_txs),
            "income": today_income,
            "expense": today_expense,
            "net": today_income - today_expense,
        },
        "account_balances": account_balances,
        "category_pnl": category_pnl,
        "monthly_trend": monthly,
        "category_donut": donut_data,
    })


# ── API: Accounts recent transactions (batch — 1 request instead of 12) ──────

@app.route("/api/accounts/recent-transactions")
@login_required
def api_accounts_recent_transactions():
    result = {}
    for acc in ACCOUNTS:
        txs = (
            Transaction.query
            .filter_by(account=acc, is_deleted=False)
            .order_by(Transaction.date.desc(), Transaction.created_at.desc())
            .limit(5).all()
        )
        result[acc] = [t.to_dict() for t in txs]
    return jsonify(result)


# ── API: Categories ────────────────────────────────────────────────────────────

@app.route("/api/categories")
@login_required
def api_categories():
    cats = Category.query.filter_by(parent_id=None, is_active=True).order_by(Category.sort_order).all()
    result = []
    for c in cats:
        subs = (
            Category.query.filter_by(parent_id=c.id, is_active=True)
            .order_by(Category.sort_order)
            .all()
        )
        result.append({
            "id": c.id,
            "name": c.name,
            "subcategories": [{"id": s.id, "name": s.name} for s in subs],
        })
    return jsonify({"categories": result})


@app.route("/api/categories", methods=["POST"])
@login_required
def api_create_category():
    data = request.get_json()
    name = (data.get("name") or "").strip()
    parent_id = data.get("parent_id")
    if not name:
        return jsonify({"error": "Name is required"}), 400
    cat = Category(name=name, parent_id=parent_id or None)
    db.session.add(cat)
    db.session.commit()
    return jsonify({"id": cat.id, "name": cat.name, "message": "Category created"}), 201


# ── API: Opening Balances ─────────────────────────────────────────────────────

@app.route("/api/opening-balances")
@login_required
def api_get_opening_balances():
    obs = {ob.account: float(ob.balance) for ob in OpeningBalance.query.all()}
    result = [{"account": a, "balance": obs.get(a, 0.0)} for a in ACCOUNTS]
    return jsonify({"opening_balances": result})


@app.route("/api/opening-balances", methods=["POST"])
@login_required
def api_save_opening_balances():
    data = request.get_json()
    balances = data.get("balances", {})
    for account, balance in balances.items():
        ob = OpeningBalance.query.filter_by(account=account).first()
        if ob:
            ob.balance = Decimal(str(balance))
            ob.updated_at = datetime.utcnow()
        else:
            db.session.add(OpeningBalance(account=account, balance=Decimal(str(balance))))
    db.session.commit()
    return jsonify({"message": "Opening balances saved"})


# ── API: Reports ──────────────────────────────────────────────────────────────

@app.route("/api/report")
@login_required
def api_report():
    from_date_str = request.args.get("from_date")
    to_date_str = request.args.get("to_date")

    try:
        from_date = date.fromisoformat(from_date_str)
        to_date = date.fromisoformat(to_date_str)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid date range"}), 400

    txs = (
        Transaction.query
        .filter(
            Transaction.is_deleted == False,
            Transaction.date >= from_date,
            Transaction.date <= to_date,
        )
        .order_by(Transaction.date.desc())
        .all()
    )

    total_income = sum(float(t.amount) for t in txs if t.type == "Income")
    total_expense = sum(float(t.amount) for t in txs if t.type == "Expense")

    # Category breakdown
    cat_map = {}
    for t in txs:
        if t.category not in cat_map:
            cat_map[t.category] = {"category": t.category, "income": 0.0, "expense": 0.0}
        if t.type == "Income":
            cat_map[t.category]["income"] += float(t.amount)
        elif t.type in ("Expense", "Transfer"):
            cat_map[t.category]["expense"] += float(t.amount)
    for v in cat_map.values():
        v["net"] = v["income"] - v["expense"]

    # Account breakdown
    acc_map = {}
    for t in txs:
        if t.account not in acc_map:
            acc_map[t.account] = {"account": t.account, "income": 0.0, "expense": 0.0}
        if t.type == "Income":
            acc_map[t.account]["income"] += float(t.amount)
        elif t.type in ("Expense", "Transfer"):
            acc_map[t.account]["expense"] += float(t.amount)

    # Top 10 by amount
    top10 = sorted(txs, key=lambda t: float(t.amount), reverse=True)[:10]

    return jsonify({
        "from_date": from_date_str,
        "to_date": to_date_str,
        "summary": {
            "total_income": total_income,
            "total_expense": total_expense,
            "net": total_income - total_expense,
            "transaction_count": len(txs),
        },
        "category_breakdown": list(cat_map.values()),
        "account_breakdown": list(acc_map.values()),
        "top_transactions": [t.to_dict() for t in top10],
        "transactions": [t.to_dict() for t in txs],
    })


# ── API: Settings ─────────────────────────────────────────────────────────────

@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    data = request.get_json()
    current_pw = data.get("current_password", "")
    new_pw = data.get("new_password", "")

    if not current_user.check_password(current_pw):
        return jsonify({"error": "Current password is incorrect"}), 400
    if len(new_pw) < 6:
        return jsonify({"error": "New password must be at least 6 characters"}), 400

    current_user.set_password(new_pw)
    current_user.updated_at = datetime.utcnow()
    log_audit("CHANGE_PASSWORD", "users", current_user.id)
    db.session.commit()
    return jsonify({"message": "Password changed successfully"})


@app.route("/api/db-status")
@login_required
def api_db_status():
    try:
        count = Transaction.query.count()
        return jsonify({"status": "connected", "transaction_count": count})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


# ── Exports ───────────────────────────────────────────────────────────────────

def _build_excel(txs, title="Finance Export"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Transactions ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Type", "Account", "Category", "Sub-Category", "Description", "Amount (₹)", "Ref"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1e2a")
        cell.alignment = Alignment(horizontal="center")

    for t in txs:
        ws.append([
            t.date.strftime("%d/%m/%Y"),
            t.type,
            t.account,
            t.category,
            t.subcat or "",
            t.description or "",
            float(t.amount),
            t.ref or "",
        ])
        ws.cell(ws.max_row, 7).number_format = '#,##0.00'

    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            12, max((len(str(cell.value or "")) for cell in col), default=10)
        )

    # ── Sheet 2: Account Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Account Summary")
    ws2.append(["Account", "Period In (₹)", "Period Out (₹)", "Current Balance (₹)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1e2a")

    for acc in ACCOUNTS:
        # Compute period activity from the passed txs (correct for both full and report exports)
        acc_txs = [t for t in txs if t.account == acc]
        transfer_in_txs = [t for t in txs if getattr(t, "to_account", None) == acc and t.type == "Transfer"]
        period_in = (sum(float(t.amount) for t in acc_txs if t.type == "Income")
                     + sum(float(t.amount) for t in transfer_in_txs))
        period_out = sum(float(t.amount) for t in acc_txs if t.type in ("Expense", "Transfer"))
        # Current balance is always all-time (opening + all transactions ever)
        balance = get_account_balance(acc)
        ws2.append([acc, period_in, period_out, balance])
        for col in [2, 3, 4]:
            ws2.cell(ws2.max_row, col).number_format = '#,##0.00'

    ws2.column_dimensions["A"].width = 30
    for col in ["B", "C", "D"]:
        ws2.column_dimensions[col].width = 18

    # ── Sheet 3: Category Summary ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Category Summary")
    ws3.append(["Category", "Total Income (₹)", "Total Expense (₹)", "Net P&L (₹)"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1e2a")

    cat_map = {}
    for t in txs:
        if t.category not in cat_map:
            cat_map[t.category] = [0.0, 0.0]
        if t.type == "Income":
            cat_map[t.category][0] += float(t.amount)
        elif t.type in ("Expense", "Transfer"):
            cat_map[t.category][1] += float(t.amount)
    for cat, (inc, exp) in cat_map.items():
        ws3.append([cat, inc, exp, inc - exp])
        for col in [2, 3, 4]:
            ws3.cell(ws3.max_row, col).number_format = '#,##0.00'

    ws3.column_dimensions["A"].width = 30
    for col in ["B", "C", "D"]:
        ws3.column_dimensions[col].width = 18

    # ── Sheet 4: Monthly Summary ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Monthly Summary")
    ws4.append(["Month", "Income (₹)", "Expense (₹)", "Net (₹)"])
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1e2a")

    month_map = {}
    for t in txs:
        m = t.month
        if m not in month_map:
            month_map[m] = [0.0, 0.0]
        if t.type == "Income":
            month_map[m][0] += float(t.amount)
        elif t.type in ("Expense", "Transfer"):
            month_map[m][1] += float(t.amount)
    for m in sorted(month_map.keys()):
        inc, exp = month_map[m]
        ws4.append([m, inc, exp, inc - exp])
        for col in [2, 3, 4]:
            ws4.cell(ws4.max_row, col).number_format = '#,##0.00'

    ws4.column_dimensions["A"].width = 15
    for col in ["B", "C", "D"]:
        ws4.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/api/export/excel")
@login_required
def export_excel():
    txs = (
        Transaction.query.filter_by(is_deleted=False)
        .order_by(Transaction.date.desc())
        .all()
    )
    buf = _build_excel(txs)
    filename = f"finance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/export/csv")
@login_required
def export_csv():
    import csv

    txs = (
        Transaction.query.filter_by(is_deleted=False)
        .order_by(Transaction.date.desc())
        .all()
    )
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Date", "Type", "Account", "Category", "Sub-Category", "Description", "Amount", "Ref"])
    for t in txs:
        w.writerow([
            t.date.strftime("%d/%m/%Y"),
            t.type, t.account, t.category,
            t.subcat or "", t.description or "",
            float(t.amount), t.ref or "",
        ])
    out.seek(0)
    buf = io.BytesIO(out.getvalue().encode("utf-8-sig"))
    return send_file(
        buf,
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"transactions_{datetime.now().strftime('%Y%m%d')}.csv",
    )


@app.route("/api/export/report")
@login_required
def export_report():
    from_date_str = request.args.get("from_date")
    to_date_str = request.args.get("to_date")
    try:
        from_date = date.fromisoformat(from_date_str)
        to_date = date.fromisoformat(to_date_str)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid date range"}), 400

    txs = (
        Transaction.query
        .filter(
            Transaction.is_deleted == False,
            Transaction.date >= from_date,
            Transaction.date <= to_date,
        )
        .order_by(Transaction.date.desc())
        .all()
    )
    buf = _build_excel(txs, title="Finance Report")
    filename = f"report_{from_date_str}_to_{to_date_str}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
